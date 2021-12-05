import sqlite3
from Actors import *
from Assests import subjects,labs,constraints, roomEntity
import pandas as pd
import os


class DataBaseBuilder:
    t = 0
    s = 0
    r = 0

    def __init__(self):
        self._conn = sqlite3.connect('TimeTable.db')
        self._cursor = self._conn.cursor()
        try :
            ids = self._cursor.execute('SELECT _id FROM teacherTable ORDER BY _id DESC LIMIT 1').fetchone()
            self.t = int(ids[0].replace('T',''))
            ids = self._cursor.execute('SELECT id FROM sectionTable ORDER BY id DESC LIMIT 1').fetchone()
            self.s = int(ids[0].replace('S',''))
            ids = self._cursor.execute('SELECT _id FROM roomTable ORDER BY _id DESC LIMIT 1').fetchone()
            self.r = int(ids[0].replace('R', ''))
        except:
            print("Table Not Created")
            t=0
            s=0
            r=0

    # def __del__(self):
    #     self._conn.close()

    # For Internal Use by developers
    def inputDetails(self):
        name = input('Name : ')
        n = int(input('Number of Subjects : '))
        subs = []
        while (n > 0):
            subs.append([input('Subject Name : '), input('Subject Credits\Load : ')])
            n -= 1
        n = int(input('number of Labs : '))
        lab = []
        while (n > 0):
            lab.append(input('lab Name : '))
            n -= 1
        return name, subs, lab

    def createTeacherTable(self):
        self._cursor.execute(
            'CREATE TABLE IF NOT EXISTS teacherTable(_id TEXT PRIMARY KEY,name TEXT,sub1 TEXT,load1 TEXT,sub2 TEXT,load2 TEXT,sub3 TEXT,load3 TEXT,sub4 TEXT,load4 TEXT,sub5 TEXT,load5 TEXT,lab1 TEXT,lab2 TEXT,lab3 TEXT)')
        self._conn.commit()

    def createSectionTable(self):
        self._cursor.execute(
            'CREATE TABLE IF NOT EXISTS sectionTable(id TEXT PRIMARY KEY,name TEXT, branch TEXT,sem TEXT,sub1 TEXT,credit1 TEXT,sub2 TEXT,credit2 TEXT,sub3 TEXT,credit3 TEXT,sub4 TEXT,credit4 TEXT,sub5 TEXT,credit5 TEXT,sub6 TEXT,credit6 TEXT,sub7 TEXT,credit7 TEXT,sub8 TEXT,credit8 TEXT,lab1 TEXT,lab2 TEXT,lab3 TEXT,lab4 TEXT,lab5 TEXT)')
        self._conn.commit()

    def createRoomTable(self):
        self._cursor.execute(
            'CREATE TABLE IF NOT EXISTS roomTable(_id TEXT PRIMARY KEY,campus TEXT,department TEXT,building TEXT,room TEXT,capacity TEXT)')
        self._conn.commit()

    def insertTeacherDetails(self, _id, name, subs=[], lab=[]):
        for i in range(5 - len(subs)):
            subs.append([None, None])
        for i in range(3 - len(lab)):
            lab.append(None)
        self._cursor.execute('INSERT INTO teacherTable VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
                              (_id, name, subs[0][0], subs[0][1], subs[1][0], subs[1][1], subs[2][0], subs[2][1],
                               subs[3][0], subs[3][1], subs[4][0], subs[4][1],
                               lab[0], lab[1], lab[2]))
        self._conn.commit()

    def insertSectionDetails(self, _id, name, subs, lab=[]):
        for i in range(8 - len(subs)):
            subs.append([None, None])
        for i in range(5 - len(lab)):
            lab.append(None)
        self._cursor.execute('INSERT INTO sectionTable VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
                              (_id, name, subs[0][0], subs[0][1], subs[1][0], subs[1][1], subs[2][0], subs[2][1],
                               subs[3][0], subs[3][1], subs[4][0], subs[4][1], subs[5][0],
                               subs[5][1], subs[6][0], subs[6][1], subs[7][0], subs[7][1], lab[0], lab[1], lab[2],
                               lab[3], lab[4]))
        self._conn.commit()

    def insertRoomDetails(self, _id, campus, department, building, room, capacity):
        self._cursor.execute('INSERT INTO roomTable VALUES(?,?,?,?,?,?)',(_id,campus,department,building,room,capacity))
        self._conn.commit()

    def insertTeacherDetailsFromPage(self, data):
        self.t = self.t + 1
        t_id = 'T' + str(self.t)
        t_name = data['teacher_name']
        t_sub1 = data['teacher_sub1']
        periods1 = data['periods1']
        t_sub2 = data['teacher_sub2']
        periods2 = data['periods2']
        t_sub3 = data['teacher_sub3']
        periods3 = data['periods3']
        t_sub4 = data['teacher_sub4']
        periods4 = data['periods4']
        t_sub5 = data['teacher_sub5']
        periods5 = data['periods5']
        t_lab1 = data['teacher_lab1']
        t_lab2 = data['teacher_lab2']
        t_lab3 = data['teacher_lab3']

        self._cursor.execute("INSERT INTO teacherTable VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", (t_id,
                                                                                                  t_name, t_sub1,
                                                                                                  periods1, t_sub2,
                                                                                                  periods2, t_sub3,
                                                                                                  periods3, t_sub4,
                                                                                                  periods4, t_sub5,
                                                                                                  periods5,
                                                                                                  t_lab1, t_lab2,
                                                                                                  t_lab3))
        self._conn.commit()
        msg = "Successfully Inserted"
        return msg

    def setNone(self, formData):
        dataStr = str(formData)
        if len(dataStr) == 0 :
            return None
        return formData

    def insertSectionDetailsFromPage(self, data):
        self.s = self.s + 1
        sec_id = 'S' + str(self.s)
        name = data['name']
        branch = data['branch']
        sem = data['sem']
        sub1 = data['sub1']
        credit1 = data['credit1']
        sub2 = data['sub2']
        self.setNone(sub2)
        credit2 = data['credit2']
        self.setNone(credit2)
        sub3 = data['sub3']
        sub3 = self.setNone(sub3)
        credit3 = data['credit3']
        credit3 = self.setNone(credit3)
        sub4 = data['sub4']
        sub4 = self.setNone(sub4)
        credit4 = data['credit4']
        credit4 = self.setNone(credit4)
        sub5 = data['sub5']
        sub5 = self.setNone(sub5)
        credit5 = data['credit5']
        credit5 = self.setNone(credit5)
        sub6 = data['sub6']
        sub6 = self.setNone(sub6)
        credit6 = data['credit6']
        credit6 = self.setNone(credit6)
        sub7 = data['sub7']
        sub7 = self.setNone(sub7)
        credit7 = data['credit7']
        credit7 = self.setNone(credit7)
        sub8 = data['sub8']
        sub8 = self.setNone(sub8)
        credit8 = data['credit8']
        credit8 = self.setNone(credit8)
        lab1 = data['lab1']
        lab1 = self.setNone(lab1)
        lab2 = data['lab2']
        lab2 = self.setNone(lab2)
        lab3 = data['lab3']
        lab3 = self.setNone(lab3)
        lab4 = data['lab4']
        lab4 = self.setNone(lab4)
        lab5 = data['lab5']
        lab5 = self.setNone(lab5)
        self._cursor.execute("INSERT INTO sectionTable VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", (sec_id,
                                                                                                                name,branch,sem,
                                                                                                                sub1,
                                                                                                                credit1,
                                                                                                                sub2,
                                                                                                                credit2,
                                                                                                                sub3,
                                                                                                                credit3,
                                                                                                                sub4,
                                                                                                                credit4,
                                                                                                                sub5,
                                                                                                                credit5,
                                                                                                                sub6,
                                                                                                                credit6,
                                                                                                                sub7,
                                                                                                                credit7,
                                                                                                                sub8,
                                                                                                                credit8,
                                                                                                                lab1,
                                                                                                                lab2,
                                                                                                                lab3,
                                                                                                                lab4,
                                                                                                                lab5))

        self._conn.commit()
        return "Successfully Inserted"

# Insert Room Details from page
    def insertRoomDetailsFromPage(self, data):
        self.r = self.r + 1
        room_id = 'R' + str(self.r)
        campus = data['campus']
        department = data['department']
        building = data['building']
        room = data['room']
        capacity = data['capacity']
        self._cursor.execute("INSERT INTO roomTable VALUES (?,?,?,?,?,?)", (room_id,campus,department,building,room,capacity))
        self._conn.commit()
        return "Successfully Inserted"

    def remove_section(self, data):
        row = data['remove_sec']
        self._cursor.execute("DELETE FROM sectionTable WHERE LOWER(id)=?", (str(row).lower(),))
        self._conn.commit()
        count = 0
        r = self._cursor.execute("SELECT * FROM sectionTable")
        row = self._cursor.fetchall()
        while row:
            oldId = row[count][0]
            newId = "S" + str(count)
            self._cursor.execute("UPDATE sectionTable SET id=? WHERE id=?", (newId, oldId))
            self._conn.commit()
            count += 1


    def remove_teacher(self, data):
        row = data['remove_t']
        self._cursor.execute("DELETE FROM teacherTable WHERE LOWER(_id)=?", (str(row).lower(),))
        self._conn.commit()
        count = 0
        r = self._cursor.execute("SELECT * FROM teacherTable")
        row = self._cursor.fetchall()
        while row:
            oldId = row[count][0]
            newId = "T" + str(count)
            self._cursor.execute("UPDATE teacherTable SET _id=? WHERE _id=?", (newId, oldId))
            self._conn.commit()
            count += 1

    def remove_room(self, data):
        row = data['remove_room']
        self._cursor.execute("DELETE FROM roomTable WHERE LOWER(_id)=?", (str(row).lower(),))
        self._conn.commit()
        count = 0
        r = self._cursor.execute("SELECT * FROM roomTable")
        row = self._cursor.fetchall()
        while row :
            oldId = row[count][0]
            newId = "R"+str(count)
            self._cursor.execute("UPDATE roomTable SET _id=? WHERE _id=?", (newId, oldId))
            self._conn.commit()
            count += 1

    def inputSections(self):
        self.createSectionTable()
        n = int(input('Number of Sections : '))
        while n > 0:
            self.s += 1
            print(self.s)
            _id = 'S' + str(self.s)
            name, subs, lab = self.inputDetails()
            try:
                self.insertSectionDetails(_id, name, subs, lab)
                n -= 1
            except Exception as e:
                print('Exception : ', str(e))


    def inputTeachers(self):
        self.createTeacherTable()
        n = int(input('Number of Teachers : '))
        while n > 0:
            self.t += 1
            _id = 'T' + str(self.t)
            name, subs, lab = self.inputDetails()
            try:
                self.insertTeacherDetails(_id, name, subs, lab)
                n -= 1
            except Exception as e:
                print('Exception : ', str(e))


    def displayDatabase(self):
        con = self._conn
        cur = self._cursor
        con.row_factory = sqlite3.Row
        cur = con.cursor()
        cur.execute("select * from teacherTable")
        rows2 = cur.fetchall();
        cur.execute("select * from sectionTable ORDER BY sem")
        rows1 = cur.fetchall();
        cur.execute("select * from roomTable ORDER BY _id")
        rows3 = cur.fetchall();
        return rows1, rows2, rows3


class timetableGenerator(DataBaseBuilder):
    def __init__(self):
        DataBaseBuilder.__init__(self)
        self.__teachers = []
        self.__sections = []
        self.__rooms = [] #to be fetched from Database
        self.__list_of_timetables = []
        self.__list_of_names = []

    def loadDB(self):
        self.__teachers = []
        self.__sections = []
        self.__rooms = []
        self._cursor.execute('SELECT * FROM teacherTable')
        teachers = [list(row) for row in self._cursor.fetchall()]
        self._cursor.execute('SELECT * FROM sectionTable ORDER BY sem')
        sections = [list(row) for row in self._cursor.fetchall()]
        self._cursor.execute('SELECT * FROM roomTable ORDER BY _id')
        rooms = [list(row) for row in self._cursor.fetchall()]

        def makeLabObj(lab):
            return labs(lab,'3')

        def makeSubObj(sub):
            sub = [subjects(subName = sub[i],subCreditsorLoads = sub[i+1]) for i in range(0,len(sub)-1,2)]
            return sub

        for section in sections:
            section = [sec for sec in section if sec is not None]
            labsL = [makeLabObj(lab) for lab in section if "lab" in str(lab).lower()]
            subsL = makeSubObj(section[4:-len(labsL)])
            self.__sections.append(sectionTimeTable(section[0],section[1],section[2],section[3],subsL,labsL))########

        for teacher in teachers:
            teacher = [te for te in teacher if te is not None]
            labsL = [makeLabObj(lab) for lab in teacher if "lab" in str(lab).lower()]
            subsL = makeSubObj(teacher[2:-len(labsL)])
            self.__teachers.append(teacherTimeTable(teacher[0], teacher[1], subsL, labsL))

        for room in rooms:
            room = [r_details for r_details in room if r_details is not None]
            self.__rooms.append(roomEntity(room[0],room[1],room[2],room[3],room[4],room[5]))

    def assignTeacherstoSections(self):
        for sidx, section in enumerate(self.__sections):
            secSubs = section.getSubjects()
            for sub in secSubs:
                for tidx, teacher in enumerate(self.__teachers):
                    tsub = teacher.getSubjects()
                    if sub.getName() in [t.getName() for t in
                                         tsub] and teacher.getCanAssignSec() and not teacher.checkSectionExistance(
                            sidx):
                        temp = [t for t in tsub if t.getName() == sub.getName()][0]
                        if teacher.getLoad(temp) != 0:
                            teacher.updateSubjectLoad(temp)
                            teacher.addSectionforSubject(sub, sidx)  # sectionAlloted[sub] = sectionIndexinList
                            teacher.incrementClassesAlloted()
                            section.addTeacherforSubject(sub, tidx)  # teacherAlloted[sub] = teacherIndexinList
                            if teacher.getNumClassesAlloted() == teacher.getMaxClassesToAllot():
                                teacher.setCanAssignSec(False)
                            break

        for sec in self.__sections:
            print(sec.getName())
            s = sec.getTeachersAlloted()
            for k, v in s.items():
                print(k.getName(), v)
            print('-' * 50)
        for teacher in self.__teachers:
            print(teacher.getName())
            s = teacher.getSectionAlloted()
            for k, v in s.items():
                print(k.getName(), v)

    def build_timeTable(self):
        for sec in self.__sections:
            sec.assignNoClasses()
            sec.assignLabs()
            sec.assignRooms(self.__rooms)

            if not sec.enoughSlots():
                print('Not enogh space.!')
                continue
            # temp_timeTable = sec.getTimeTable()
            temp_subs = sec.getSubjects()
            count = 0
            slots = [0, 3, 6]
            days = np.random.choice(range(5), 5, replace=False)
            for day in days:
                validSlots = sec.getValidSlots(day)
                for v in validSlots:
                    attempts = 0
                    while not sec.isfull(day, v) and attempts < 10 and len(temp_subs) != 0:
                        attempts += 1
                        if str.encode(str(temp_subs[count].getName())) not in sec.getTimeTableforDay(day):
                            teacher = self.__teachers[sec.getTeacherAllotedforSubject(temp_subs[count])]
                            if teacher.check(day, slots[v] + sec.getFilled(day, v), sec.getName()):
                                # print('t',day,v,count,temp_subs[count].getName())
                                sec.updateTimeTable(day, slots[v] + sec.getFilled(day, v),
                                                    temp_subs[count].getName())
                                sec.decreaseCredits(temp_subs[count])
                                sec.incrementAllSlotsby(day, slots[v] // 3, 1)
                                attempts = 0
                                if sec.getCredits(temp_subs[count]) == 0:
                                    temp_subs.pop(count).getName()
                        if len(temp_subs) != 0:
                            count = (count + 1) % len(temp_subs)

            if len(temp_subs) != 0:
                return False

            sec.addRooms(self.__rooms)

        return True

    def printDetails(self):
        for sec in self.__sections:
            print(sec.getTimeTable())
            print()
        print('--' * 50)
        for teacher in self.__teachers:
            print(teacher.getTimeTable())
            print()

    def generateExcel(self):
        if os.path.isfile('SectionsTable.xlsx'):
            os.remove("SectionsTable.xlsx")
        if os.path.isfile('TeacherTable.xlsx'):
            os.remove("TeacherTable.xlsx")

        days = ['MON', 'TUE', 'WED', 'THURS', 'FRI']
        period = ['room', '8-9', '9-10', '10-11', 'room', '11-12', '12-1', '1-2', 'room', '3-4', '4-5', '5-6']
        for section in self.__sections:
            time_table_sec_formatted = pd.DataFrame(data=section.getTimeTable(), index=days, columns=period,
                                                    dtype='U20')
            for time in period:
                time_table_sec_formatted[time] = time_table_sec_formatted[time].replace([''], '-')
            # print(time_table_sec_formatted)

            self.__list_of_timetables.append(time_table_sec_formatted)
            self.__list_of_names.append(section.getName() + '(' + section.getSemester() + ')')
            if not os.path.isfile('SectionsTable.xlsx'):
                time_table_sec_formatted.to_excel('SectionsTable.xlsx',
                                                  sheet_name=section.getName() + '(' + section.getSemester() + ')')
            else:
                from openpyxl import load_workbook
                book = load_workbook('SectionsTable.xlsx')
                writer = pd.ExcelWriter('SectionsTable.xlsx', engine='openpyxl')
                writer.book = book
                time_table_sec_formatted.to_excel(writer,
                                                  sheet_name=section.getName() + '(' + section.getSemester() + ')')
                writer.save()
                writer.close()
        period = ['8-9', '9-10', '10-11', '11-12', '12-1', '1-2', '3-4', '4-5', '5-6']
        for teacher in self.__teachers:
            time_table_sec_formatted = pd.DataFrame(data=teacher.getTimeTable(), index=days, columns=period,
                                                    dtype='U20')
            for time in period:
                time_table_sec_formatted[time] = time_table_sec_formatted[time].replace([''], '-')
            # print(time_table_sec_formatted)
            self.__list_of_timetables.append(time_table_sec_formatted)
            self.__list_of_names.append(teacher.getName())

            if not os.path.isfile('TeacherTable.xlsx'):
                time_table_sec_formatted.to_excel('TeacherTable.xlsx', sheet_name=teacher.getName())
            else:
                from openpyxl import load_workbook
                book = load_workbook('TeacherTable.xlsx')
                writer = pd.ExcelWriter('TeacherTable.xlsx', engine='openpyxl')
                writer.book = book
                time_table_sec_formatted.to_excel(writer, sheet_name=teacher.getName())
                writer.save()
                writer.close()
        return list(zip(self.__list_of_names, self.__list_of_timetables))


    if __name__ == '__main__':
        pass
        # #dbController = DataBaseBuilder()
        # #dbController.inputTeachers()
        # #dbController.inputSections()
        # while(True):
        #     controller = timetableGenerator()
        #     controller.assignTeacherstoSections()
        #     if(controller.build_timeTable()):
        #         #controller.printDetails()
        #         controller.generateExcel()
        #         break
